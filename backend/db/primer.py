import psycopg2
from openpyxl import load_workbook

from os import listdir
from os.path import isfile, join

conn = psycopg2.connect(host="localhost",database="postgres", user="postgres", password="daddyluke123")
conn.autocommit = True

cursor = conn.cursor()

# Print PostgreSQL Connection properties
print ( conn.get_dsn_parameters(),"\n")

# Print PostgreSQL version
cursor.execute("SELECT version();")
record = cursor.fetchone()
print("You are connected to - ", record,"\n")

drop_table_query = "DROP TABLE Prices"

# try:
#     cursor.execute(drop_table_query)
#     print('dropped prices table')
# except:
#     print("prices table didn't exist, continuing...")

# create_table_query = '''
# CREATE TABLE Prices(
#     ticker VARCHAR(10) NOT NULL,
#     date DATE NOT NULL,
#     open DECIMAL,
#     high DECIMAL,
#     low DECIMAL,
#     close DECIMAL,
#     volume BIGINT,
#     smavg BIGINT,
#     PRIMARY KEY (ticker,date)
#     ); '''

# cursor.execute(create_table_query)

pathToData = "../../algDev/data/"
equitiesPath = pathToData + 'equities/'

postgres_insert_query = """ INSERT INTO Prices (ticker, date, open, high, low, close, volume, smavg) VALUES ('{}','{}',{},{},{},{},{},{})"""


onlyFiles = [f for f in listdir(equitiesPath) if isfile(join(equitiesPath, f)) and f[-5:] == ".xlsx"]
print('files?')
onlyFiles.sort()
print(onlyFiles)

# completed = ['AAPL.xlsx', 'ABBV.xlsx', 'ABT.xlsx', 'ACN.xlsx', 'ADBE.xlsx', 'ADP.xlsx', 'AMGN.xlsx', 'AMT.xlsx', 'AMZN.xlsx', 'ANTM.xlsx', 'AVGO.xlsx', 'AXP.xlsx', 'BA.xlsx', 'BAC.xlsx', 'BDX.xlsx', 'BKNG.xlsx', 'BLK.xlsx', 'BMY.xlsx', 'BRK.B.xlsx', 'C.xlsx', 'CAT.xlsx', 'CB.xlsx', 'CHTR.xlsx', 'CI.xlsx', 'CL.xlsx', 'CMCSA.xlsx', 'CME.xlsx', 'COST.xlsx', 'CRM.xlsx', 'CSCO.xlsx', 'CVS.xlsx', 'CVX.xlsx', 'D.xlsx', 'DHR.xlsx', 'DIS.xlsx', 'DUK.xlsx', 'FB.xlsx', 'FIS.xlsx', 'FISV.xlsx', 'GE.xlsx', 'GILD.xlsx', 'GOOG.xlsx', 'GOOGL.xlsx', 'GS.xlsx', 'HD.xlsx', 'HON.xlsx', 'IBM.xlsx', 'INTU.xlsx', 'ISRG.xlsx', 'JNJ.xlsx', 'JPM.xlsx', 'KO.xlsx', 'LIN.xlsx', 'LLY.xlsx', 'LMT.xlsx', 'LOW.xlsx', 'MA.xlsx', 'MCD.xlsx', 'MDLZ.xlsx', 'MDT.xlsx', 'MMM.xlsx', 'MO.xlsx', 'MRK.xlsx', 'MS.xlsx', 'MSFT.xlsx', 'NEE.xlsx']
completedSet = set(completed)

for file in onlyFiles:
    if file in completedSet:
        continue
    tickerName = file[:-5]

    wb = load_workbook(equitiesPath + file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if not row[0]:
                break
            
            dateObj = row[0]
            dateStr = str(dateObj.year) + '-' + str(dateObj.month) + '-' + str(dateObj.day)

            toUse = (tickerName, dateStr, row[1], row[2], row[3], row[4], row[5], row[6])

            modified = False
            lst = list(toUse)
            for i in range(len(toUse)):
                if not toUse[i]:
                    modified = True
                    lst[i] = "NULL"
            if modified:
                toUse = tuple(lst)

            formatted = postgres_insert_query.format(*toUse)
            print(formatted)
            cursor.execute(formatted)

        except (Exception, psycopg2.Error) as error :
            print ("Error while connecting to PostgreSQL", error)


if(conn):
    cursor.close()
    conn.close()
    print("PostgreSQL connection is closed")